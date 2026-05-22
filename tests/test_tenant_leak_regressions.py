from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO

import pytest

from backend.models import (
    AdminUser,
    Assignment,
    Intervenant,
    NotificationJob,
    Request,
    Structure,
    User,
)
from backend.helpchain_backend.src.models import Case

pytestmark = pytest.mark.spine


def _login_admin(client, app, admin_user: AdminUser) -> None:
    with client.session_transaction() as sess:
        sess["_user_id"] = str(admin_user.id)
        sess["user_id"] = admin_user.id
        sess["admin_id"] = admin_user.id
        sess["admin_user_id"] = admin_user.id
        sess["role"] = admin_user.role
        sess["is_authenticated"] = True
        sess["is_admin"] = True
        sess["admin_logged_in"] = True
        sess[app.config.get("MFA_SESSION_KEY", "mfa_ok")] = True
        sess["mfa_required"] = True
        sess["mfa_ok_until"] = (
            datetime.now(UTC) + timedelta(minutes=30)
        ).isoformat()
        sess["admin_mfa_last_verified"] = 4102444800
        sess["admin_mfa_user_id"] = admin_user.id


def _make_structure(session, *, name: str, slug: str) -> Structure:
    row = Structure(name=name, slug=slug)
    session.add(row)
    session.flush()
    return row


def _make_admin(
    session,
    *,
    username: str,
    email: str,
    role: str,
    structure_id: int | None,
) -> AdminUser:
    row = AdminUser(
        username=username,
        email=email,
        password_hash="x",
        role=role,
        structure_id=structure_id,
        is_active=True,
        mfa_enabled=True,
        totp_secret="tenant-leak-test",
    )
    session.add(row)
    session.flush()
    return row


def _make_user(session, *, username: str, email: str, structure_id: int | None = None) -> User:
    row = User(
        username=username,
        email=email,
        password_hash="x",
        role="requester",
        is_active=True,
        structure_id=structure_id,
    )
    session.add(row)
    session.flush()
    return row


def _make_request(
    session,
    *,
    title: str,
    user_id: int,
    structure_id: int,
    status: str = "open",
    owner_id: int | None = None,
    priority: str = "normal",
    created_at: datetime | None = None,
    updated_at: datetime | None = None,
) -> Request:
    row = Request(
        title=title,
        description=f"Description for {title}",
        category="general",
        user_id=user_id,
        structure_id=structure_id,
        status=status,
        owner_id=owner_id,
        priority=priority,
        created_at=created_at or datetime.now(UTC).replace(tzinfo=None),
        updated_at=updated_at or created_at or datetime.now(UTC).replace(tzinfo=None),
    )
    session.add(row)
    session.flush()
    return row


def _make_case(
    session,
    *,
    request_id: int,
    structure_id: int,
    status: str = "new",
    owner_user_id: int | None = None,
    priority: str = "normal",
    risk_score: int = 0,
    last_activity_at: datetime | None = None,
) -> Case:
    row = Case(
        request_id=request_id,
        structure_id=structure_id,
        status=status,
        owner_user_id=owner_user_id,
        priority=priority,
        risk_score=risk_score,
        last_activity_at=last_activity_at,
        created_at=datetime.now(UTC).replace(tzinfo=None),
        updated_at=datetime.now(UTC).replace(tzinfo=None),
    )
    session.add(row)
    session.flush()
    return row


def _seed_scoped_admin(session, *, prefix: str = "tenant_scope"):
    structure_a = _make_structure(
        session,
        name=f"{prefix} Alpha",
        slug=f"{prefix}-alpha",
    )
    structure_b = _make_structure(
        session,
        name=f"{prefix} Beta",
        slug=f"{prefix}-beta",
    )
    admin = _make_admin(
        session,
        username=f"{prefix}_admin",
        email=f"{prefix}_admin@test.local",
        role="superadmin",
        structure_id=structure_a.id,
    )
    user_a = _make_user(
        session,
        username=f"{prefix}_user_a",
        email=f"{prefix}_user_a@test.local",
        structure_id=structure_a.id,
    )
    user_b = _make_user(
        session,
        username=f"{prefix}_user_b",
        email=f"{prefix}_user_b@test.local",
        structure_id=structure_b.id,
    )
    session.commit()
    return structure_a, structure_b, admin, user_a, user_b


def test_cross_tenant_request_visibility_export_and_detail_are_scoped(app, session):
    structure_a, structure_b, admin, user_a, user_b = _seed_scoped_admin(
        session, prefix="request_scope"
    )
    visible = _make_request(
        session,
        title="tenant-a-request-visible",
        user_id=user_a.id,
        structure_id=structure_a.id,
        status="pending",
    )
    hidden = _make_request(
        session,
        title="tenant-b-request-hidden",
        user_id=user_b.id,
        structure_id=structure_b.id,
        status="pending",
    )
    session.commit()

    client = app.test_client()
    _login_admin(client, app, admin)

    listing = client.get("/admin/requests")
    assert listing.status_code == 200
    listing_html = listing.get_data(as_text=True)
    assert "tenant-a-request-visible" in listing_html
    assert "tenant-b-request-hidden" not in listing_html

    own_detail = client.get(f"/admin/requests/{visible.id}")
    assert own_detail.status_code == 200
    assert "tenant-a-request-visible" in own_detail.get_data(as_text=True)

    other_detail = client.get(f"/admin/requests/{hidden.id}")
    assert other_detail.status_code == 404

    export_csv = client.get("/admin/requests/export.csv")
    assert export_csv.status_code == 200
    export_text = export_csv.get_data(as_text=True)
    assert "tenant-a-request-visible" in export_text
    assert "tenant-b-request-hidden" not in export_text


def test_cross_tenant_request_export_xlsx_matches_listing_scope(app, session):
    from openpyxl import load_workbook

    structure_a, structure_b, admin, user_a, user_b = _seed_scoped_admin(
        session, prefix="request_xlsx_scope"
    )
    _make_request(
        session,
        title="tenant-a-xlsx-visible",
        user_id=user_a.id,
        structure_id=structure_a.id,
        status="pending",
    )
    _make_request(
        session,
        title="tenant-b-xlsx-hidden",
        user_id=user_b.id,
        structure_id=structure_b.id,
        status="pending",
    )
    session.commit()

    client = app.test_client()
    _login_admin(client, app, admin)

    listing = client.get("/admin/requests")
    assert listing.status_code == 200
    listing_html = listing.get_data(as_text=True)
    assert "tenant-a-xlsx-visible" in listing_html
    assert "tenant-b-xlsx-hidden" not in listing_html

    export_xlsx = client.get("/admin/requests/export.xlsx")
    assert export_xlsx.status_code == 200

    workbook = load_workbook(BytesIO(export_xlsx.get_data()))
    sheet = workbook.active
    values = {
        str(cell)
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    }
    assert "tenant-a-xlsx-visible" in values
    assert "tenant-b-xlsx-hidden" not in values


def test_cross_tenant_legacy_api_export_is_scoped(app, session):
    from backend.helpchain_backend.src.jwt_utils import encode_access_token
    from openpyxl import load_workbook

    structure_a, structure_b, admin, user_a, user_b = _seed_scoped_admin(
        session, prefix="legacy_api_export_scope"
    )
    _make_request(
        session,
        title="tenant-a-api-export-visible",
        user_id=user_a.id,
        structure_id=structure_a.id,
        status="pending",
    )
    _make_request(
        session,
        title="tenant-b-api-export-hidden",
        user_id=user_b.id,
        structure_id=structure_b.id,
        status="pending",
    )
    session.commit()

    with app.app_context():
        token = encode_access_token(admin.id)

    client = app.test_client()
    response = client.get(
        "/api/export?format=excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.get_data()))
    sheet = workbook.active
    values = {
        str(cell)
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    }
    assert "tenant-a-api-export-visible" in values
    assert "tenant-b-api-export-hidden" not in values


def test_structure_scoped_admin_cannot_use_global_volunteer_export(app, session):
    _structure_a, _structure_b, admin, _user_a, _user_b = _seed_scoped_admin(
        session, prefix="volunteer_export_scope"
    )
    client = app.test_client()
    _login_admin(client, app, admin)

    response = client.get("/admin/export_volunteers")
    assert response.status_code == 403


def test_cross_tenant_case_visibility_and_mutation_are_scoped(app, session):
    structure_a, structure_b, admin, user_a, user_b = _seed_scoped_admin(
        session, prefix="case_scope"
    )
    visible_request = _make_request(
        session,
        title="tenant-a-case-visible",
        user_id=user_a.id,
        structure_id=structure_a.id,
        status="open",
    )
    hidden_request = _make_request(
        session,
        title="tenant-b-case-hidden",
        user_id=user_b.id,
        structure_id=structure_b.id,
        status="open",
    )
    visible_case = _make_case(
        session,
        request_id=visible_request.id,
        structure_id=structure_a.id,
        priority="critical",
        risk_score=92,
    )
    hidden_case = _make_case(
        session,
        request_id=hidden_request.id,
        structure_id=structure_b.id,
        priority="critical",
        risk_score=95,
    )
    hidden_owner = _make_admin(
        session,
        username="case_scope_hidden_owner",
        email="case_scope_hidden_owner@test.local",
        role="admin",
        structure_id=structure_b.id,
    )
    session.commit()

    client = app.test_client()
    _login_admin(client, app, admin)

    listing = client.get("/admin/cases")
    assert listing.status_code == 200
    listing_html = listing.get_data(as_text=True)
    assert "tenant-a-case-visible" in listing_html
    assert "tenant-b-case-hidden" not in listing_html

    own_detail = client.get(f"/admin/cases/{visible_case.id}")
    assert own_detail.status_code == 200
    assert "tenant-a-case-visible" in own_detail.get_data(as_text=True)

    other_detail = client.get(f"/admin/cases/{hidden_case.id}")
    assert other_detail.status_code == 404

    blocked = client.post(
        f"/admin/cases/{hidden_case.id}/assign-owner",
        data={"owner_user_id": str(hidden_owner.id)},
        follow_redirects=False,
    )
    assert blocked.status_code == 404

    session.expire_all()
    refreshed = session.get(Case, hidden_case.id)
    assert refreshed is not None
    assert refreshed.owner_user_id is None


def test_cross_tenant_assignment_visibility_is_scoped_on_intervenant_detail(app, session):
    structure_a, structure_b, admin, user_a, user_b = _seed_scoped_admin(
        session, prefix="assignment_scope"
    )
    intervenant_a = Intervenant(
        structure_id=structure_a.id,
        name="Intervenant Alpha",
        actor_type="social_worker",
        email="intervenant-alpha@test.local",
        availability="available",
        is_active=True,
    )
    intervenant_b = Intervenant(
        structure_id=structure_b.id,
        name="Intervenant Beta",
        actor_type="social_worker",
        email="intervenant-beta@test.local",
        availability="available",
        is_active=True,
    )
    session.add_all([intervenant_a, intervenant_b])
    session.flush()

    request_a = _make_request(
        session,
        title="tenant-a-assignment-visible",
        user_id=user_a.id,
        structure_id=structure_a.id,
        status="new",
    )
    request_b = _make_request(
        session,
        title="tenant-b-assignment-hidden",
        user_id=user_b.id,
        structure_id=structure_b.id,
        status="new",
    )
    session.add_all(
        [
            Assignment(
                request_id=request_a.id,
                intervenant_id=intervenant_a.id,
                structure_id=structure_a.id,
                assigned_by_admin_id=admin.id,
                status="active",
            ),
            Assignment(
                request_id=request_b.id,
                intervenant_id=intervenant_b.id,
                structure_id=structure_b.id,
                assigned_by_admin_id=admin.id,
                status="active",
            ),
        ]
    )
    session.commit()

    client = app.test_client()
    _login_admin(client, app, admin)

    detail = client.get(
        f"/admin/structures/{structure_a.id}/intervenants/{intervenant_a.id}"
    )
    assert detail.status_code == 200
    detail_html = detail.get_data(as_text=True)
    assert "tenant-a-assignment-visible" in detail_html
    assert "tenant-b-assignment-hidden" not in detail_html

    cross_tenant_detail = client.get(
        f"/admin/structures/{structure_b.id}/intervenants/{intervenant_b.id}"
    )
    assert cross_tenant_detail.status_code == 403


def test_cross_tenant_notification_visibility_and_retry_are_scoped(app, session):
    structure_a, structure_b, admin, _user_a, _user_b = _seed_scoped_admin(
        session, prefix="notification_scope"
    )
    visible_job = NotificationJob(
        channel="email",
        event_type="tenant_scope_visible",
        recipient="tenant-a-visible@test.local",
        status="failed",
        structure_id=structure_a.id,
    )
    hidden_job = NotificationJob(
        channel="email",
        event_type="tenant_scope_hidden",
        recipient="tenant-b-hidden@test.local",
        status="failed",
        structure_id=structure_b.id,
    )
    session.add_all([visible_job, hidden_job])
    session.commit()

    client = app.test_client()
    _login_admin(client, app, admin)

    listing = client.get("/ops/notifications?status=failed")
    assert listing.status_code == 200
    listing_html = listing.get_data(as_text=True)
    assert "tenant-a-visible@test.local" in listing_html
    assert "tenant-b-hidden@test.local" not in listing_html

    retry_hidden = client.post(
        f"/ops/notifications/{hidden_job.id}/retry",
        data={},
        follow_redirects=True,
    )
    assert retry_hidden.status_code == 200
    assert "Notification introuvable." in retry_hidden.get_data(as_text=True)

    retry_visible = client.post(
        f"/ops/notifications/{visible_job.id}/retry",
        data={},
        follow_redirects=True,
    )
    assert retry_visible.status_code == 200

    session.expire_all()
    refreshed_visible = session.get(NotificationJob, visible_job.id)
    refreshed_hidden = session.get(NotificationJob, hidden_job.id)
    assert refreshed_visible is not None
    assert refreshed_hidden is not None
    assert refreshed_visible.status in {"pending", "done", "sent", "failed", "dead_letter"}
    assert refreshed_hidden.status == "failed"


def test_cross_tenant_dashboard_kpis_and_ops_workspace_exclude_other_tenant_data(
    app, session
):
    structure_a, structure_b, admin, user_a, user_b = _seed_scoped_admin(
        session, prefix="kpi_scope"
    )
    now = datetime.now(UTC).replace(tzinfo=None, microsecond=0)
    stale_time = now - timedelta(hours=80)

    visible_dashboard = _make_request(
        session,
        title="tenant-a-dashboard-visible",
        user_id=user_a.id,
        structure_id=structure_a.id,
        status="open",
        owner_id=None,
        priority="high",
        created_at=now - timedelta(days=8),
        updated_at=stale_time,
    )
    _make_request(
        session,
        title="tenant-b-dashboard-hidden",
        user_id=user_b.id,
        structure_id=structure_b.id,
        status="open",
        owner_id=None,
        priority="high",
        created_at=now - timedelta(days=8),
        updated_at=stale_time,
    )
    _make_case(
        session,
        request_id=visible_dashboard.id,
        structure_id=structure_a.id,
        priority="critical",
        risk_score=95,
        last_activity_at=stale_time,
    )
    hidden_request = (
        session.query(Request).filter_by(title="tenant-b-dashboard-hidden").first()
    )
    assert hidden_request is not None
    _make_case(
        session,
        request_id=hidden_request.id,
        structure_id=structure_b.id,
        priority="critical",
        risk_score=95,
        last_activity_at=stale_time,
    )
    session.add_all(
        [
            NotificationJob(
                channel="email",
                event_type="kpi-visible-failed",
                recipient="visible-failed@test.local",
                status="failed",
                structure_id=structure_a.id,
            ),
            NotificationJob(
                channel="email",
                event_type="kpi-hidden-failed",
                recipient="hidden-failed@test.local",
                status="failed",
                structure_id=structure_b.id,
            ),
        ]
    )
    session.commit()

    client = app.test_client()
    _login_admin(client, app, admin)

    dashboard = client.get("/admin/api/dashboard?days=30")
    assert dashboard.status_code == 200
    dashboard_payload = dashboard.get_json()
    assert dashboard_payload["total_requests"] == 1
    assert dashboard_payload["counts_by_status"] == {"open": 1}

    ops_kpis = client.get("/admin/api/ops-kpis?days=30")
    assert ops_kpis.status_code == 200
    ops_payload = ops_kpis.get_json()
    assert ops_payload["new_requests"] == 1
    assert ops_payload["resolved_requests"] == 0
    assert ops_payload["stale_over_7d"] == 1

    workspace = client.get("/ops/workspace", follow_redirects=False)
    assert workspace.status_code == 200
    workspace_html = workspace.get_data(as_text=True)
    assert "tenant-a-dashboard-visible" in workspace_html
    assert "tenant-b-dashboard-hidden" not in workspace_html
    assert "hidden-failed@test.local" not in workspace_html


def test_cross_tenant_operational_report_export_is_scoped(app, session):
    from openpyxl import load_workbook

    structure_a, structure_b, admin, user_a, user_b = _seed_scoped_admin(
        session, prefix="report_scope"
    )
    now = datetime.now(UTC).replace(tzinfo=None)
    _make_request(
        session,
        title="tenant-a-report-visible",
        user_id=user_a.id,
        structure_id=structure_a.id,
        status="open",
        priority="critical",
        created_at=now - timedelta(days=1),
    )
    _make_request(
        session,
        title="tenant-b-report-hidden",
        user_id=user_b.id,
        structure_id=structure_b.id,
        status="open",
        priority="critical",
        created_at=now - timedelta(days=1),
    )
    session.commit()

    client = app.test_client()
    _login_admin(client, app, admin)

    export_xlsx = client.get("/admin/reports/operations/export.xlsx?days=7")
    assert export_xlsx.status_code == 200

    workbook = load_workbook(BytesIO(export_xlsx.get_data()))
    sheet = next(
        worksheet
        for worksheet in workbook.worksheets
        if "Situations" in worksheet.title
    )
    values = {
        str(cell)
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    }

    assert "tenant-a-report-visible" in values
    assert "tenant-b-report-hidden" not in values
